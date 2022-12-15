from contextlib import nullcontext
from sre_parse import State
import openpyxl, os
from openpyxl import Workbook, load_workbook

# Changes to the directory of the excel file
os.chdir('/Users/yasintraiba/Desktop/Automation/namingExcel')
# loads the excel file into the variable wb
wb = load_workbook('project_data.xlsx')

# Creates a variable for each sheet in the excel file
ws1 = wb['popups']
ws2 = wb['rawData']

states = ["AL", "AK", "AR", "AZ", "AR", "CA", 'CO','CT',"DE", "FL", "GA", "HI", "ID", "IL", 
        "IN", "IA", "KS", "KY", "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", 
        "NV", "NH", "NJ", "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC", "SD",    
        "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY", "DC", "Abroad"]
# Loops through the rawData sheet and adds the data to the popups sheet
for row in ws2.iter_rows(min_row=1, min_col=3, max_col=7, values_only=True):
    for state in states:
        if row[1] == state:
            ws1.cell(row=states.index(state), column=5).value = int(ws1.cell(states.index(state), column=5).value) + 1    
            if row[3] !=  None and row[3] != ' ':
                ws1.cell(states.index(state), column=6).value = int(ws1.cell(states.index(state), column=6).value) + int(row[3])
            if row[4] !=  None and row[4] != ' ':
                ws1.cell(states.index(state), column=7).value = ws1.cell(states.index(state), column=7).value + int(row[4])
            if row[0] == 'Pavement Structure Evaluation':
                ws1.cell(states.index(state), column=8).value = int(ws1.cell(states.index(state), column=8).value) + 1
            elif row[0] == 'Bridge Deck Scanning':
                ws1.cell(states.index(state), column=9).value = int(ws1.cell(states.index(state), column=9).value) + 1
            elif row[0] == 'Other Projects':
                ws1.cell(states.index(state), column=10).value = int(ws1.cell(states.index(state), column=10).value) + 1  
            elif row[0] == 'Concrete Structure Scanning':
                ws1.cell(states.index(state), column=11).value = int(ws1.cell(states.index(state), column=11).value) + 1
            elif row[0] == 'Consulting & Research':
                ws1.cell(states.index(state), column=12).value = int(ws1.cell(states.index(state), column=12).value) + 1
            elif row[0] == 'Subsurface Utility Engineering':
                ws1.cell(states.index(state), column=13).value = int(ws1.cell(states.index(state), column=13).value) + 1
        

wb.save("/Users/yasintraiba/Desktop/Infrasense-map/python/project_data.xlsx")

